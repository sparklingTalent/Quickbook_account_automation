"""Export reports to various formats."""
import pandas as pd
from pathlib import Path
from typing import Optional, List, Dict
from google.oauth2.credentials import Credentials
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, Alignment, PatternFill
from config import settings


class ReportExporter:
    """Export reports to Excel, Google Sheets, etc."""
    
    def __init__(self):
        """Initialize exporter."""
        self.sheets_service = None
        if settings.google_sheets_credentials_path:
            self._init_google_sheets()
    
    def _init_google_sheets(self):
        """Initialize Google Sheets API client."""
        try:
            creds = service_account.Credentials.from_service_account_file(
                settings.google_sheets_credentials_path,
                scopes=['https://www.googleapis.com/auth/spreadsheets']
            )
            self.sheets_service = build('sheets', 'v4', credentials=creds)
        except Exception as e:
            print(f"Warning: Could not initialize Google Sheets: {e}")
    
    def export_to_excel(self, df: pd.DataFrame, filepath: str, 
                       department_data: Optional[pd.DataFrame] = None,
                       trends_data: Optional[pd.DataFrame] = None,
                       include_charts: bool = True,
                       view_mode: str = "historical"):
        """
        Export DataFrame to Excel file with optional charts.
        
        Args:
            df: Main variance report DataFrame
            filepath: Path to save Excel file
            department_data: Department breakdown data for charts
            trends_data: Historical trends data for charts
            include_charts: Whether to include charts in the Excel file
        """
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        workbook = Workbook()
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Create main data sheet
        ws_data = workbook.create_sheet('Variance Report')
        
        # Write header
        for c_idx, col_name in enumerate(df.columns, start=1):
            cell = ws_data.cell(row=1, column=c_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Write data rows
        for r_idx, row in enumerate(df.itertuples(index=False), start=1):
            for c_idx, value in enumerate(row, start=1):
                ws_data.cell(row=r_idx + 1, column=c_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws_data.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_data.column_dimensions[column_letter].width = adjusted_width
        
        # Add charts if requested and data available
        if include_charts:
            # Department Breakdown Chart (always included)
            if department_data is not None and not department_data.empty:
                self._add_department_chart(workbook, department_data)
            
            # Historical Trends Chart (only for historical view)
            if view_mode == "historical" and trends_data is not None and not trends_data.empty:
                self._add_trends_chart(workbook, trends_data)
                # Add Monthly Comparison chart for historical view
                self._add_monthly_comparison_chart(workbook, trends_data)
            
            # Employee Variance Chart (always included)
            self._add_employee_chart(workbook, df)
        
        workbook.save(filepath)
        return filepath
    
    def _add_department_chart(self, workbook: Workbook, dept_df: pd.DataFrame):
        """Add department breakdown bar chart."""
        try:
            ws_chart = workbook.create_sheet('Department Charts')
            
            # Prepare department data (filter department totals)
            dept_rows = []
            for _, row in dept_df.iterrows():
                if 'DEPARTMENT TOTAL' in str(row.get('Employee Name', '')):
                    dept_name = row.get('Department', 'Unknown')
                    dept_rows.append({
                        'Department': dept_name,
                        'Budget': row.get('Budget', 0),
                        'Actual': row.get('Actual', 0),
                        'Variance': row.get('Variance', 0)
                    })
            
            if not dept_rows:
                return
            
            dept_data = pd.DataFrame(dept_rows)
            
            # Write department data
            headers = ['Department', 'Budget', 'Actual', 'Variance']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_chart.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
            
            for row_idx, (_, row) in enumerate(dept_data.iterrows(), start=2):
                ws_chart.cell(row=row_idx, column=1, value=row['Department'])
                ws_chart.cell(row=row_idx, column=2, value=row['Budget'])
                ws_chart.cell(row=row_idx, column=3, value=row['Actual'])
                ws_chart.cell(row=row_idx, column=4, value=row['Variance'])
            
            # Create bar chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Department Budget vs Actual"
            chart.y_axis.title = "Amount ($)"
            chart.x_axis.title = "Department"
            
            data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(dept_data) + 1, max_col=3)
            cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(dept_data) + 1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws_chart.add_chart(chart, "F2")
            
            # Auto-adjust column widths
            for col in range(1, 5):
                ws_chart.column_dimensions[chr(64 + col)].width = 15
        except Exception as e:
            print(f"Warning: Could not add department chart: {e}")
    
    def _add_trends_chart(self, workbook: Workbook, trends_df: pd.DataFrame):
        """Add historical trends line chart."""
        try:
            ws_chart = workbook.create_sheet('Historical Trends')
            
            # Write trends data
            headers = ['Month', 'Total Budget', 'Total Actual', 'Total Variance']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_chart.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
            
            for row_idx, (_, row) in enumerate(trends_df.iterrows(), start=2):
                ws_chart.cell(row=row_idx, column=1, value=row.get('Month', ''))
                ws_chart.cell(row=row_idx, column=2, value=row.get('Total Budget', 0))
                ws_chart.cell(row=row_idx, column=3, value=row.get('Total Actual', 0))
                ws_chart.cell(row=row_idx, column=4, value=row.get('Total Variance', 0))
            
            # Create line chart
            chart = LineChart()
            chart.title = "Historical Variance Trends"
            chart.style = 13
            chart.y_axis.title = "Amount ($)"
            chart.x_axis.title = "Month"
            
            data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(trends_df) + 1, max_col=3)
            cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(trends_df) + 1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws_chart.add_chart(chart, "F2")
            
            # Auto-adjust column widths
            for col in range(1, 5):
                ws_chart.column_dimensions[chr(64 + col)].width = 20
        except Exception as e:
            print(f"Warning: Could not add trends chart: {e}")
    
    def _add_employee_chart(self, workbook: Workbook, df: pd.DataFrame):
        """Add employee variance bar chart."""
        try:
            ws_chart = workbook.create_sheet('Employee Charts')
            
            # Filter employee rows (exclude department totals)
            employee_rows = []
            for _, row in df.iterrows():
                emp_name = row.get('Employee Name', '')
                if emp_name and 'DEPARTMENT TOTAL' not in str(emp_name) and row.get('Employee ID'):
                    employee_rows.append({
                        'Employee': emp_name[:20],  # Truncate long names
                        'Budget': row.get('Budget', 0),
                        'Actual': row.get('Actual', 0),
                        'Variance': row.get('Variance', 0)
                    })
            
            if not employee_rows:
                return
            
            emp_data = pd.DataFrame(employee_rows)
            
            # Limit to top 10 employees by variance for chart readability
            emp_data = emp_data.nlargest(10, 'Variance', keep='all')
            
            # Write employee data
            headers = ['Employee', 'Budget', 'Actual', 'Variance']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_chart.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
            
            for row_idx, (_, row) in enumerate(emp_data.iterrows(), start=2):
                ws_chart.cell(row=row_idx, column=1, value=row['Employee'])
                ws_chart.cell(row=row_idx, column=2, value=row['Budget'])
                ws_chart.cell(row=row_idx, column=3, value=row['Actual'])
                ws_chart.cell(row=row_idx, column=4, value=row['Variance'])
            
            # Create bar chart
            chart = BarChart()
            chart.type = "bar"
            chart.style = 10
            chart.title = "Top 10 Employees - Budget vs Actual"
            chart.y_axis.title = "Employee"
            chart.x_axis.title = "Amount ($)"
            
            data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(emp_data) + 1, max_col=3)
            cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(emp_data) + 1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws_chart.add_chart(chart, "F2")
            
            # Auto-adjust column widths
            ws_chart.column_dimensions['A'].width = 25
            for col in range(2, 5):
                ws_chart.column_dimensions[chr(64 + col)].width = 15
        except Exception as e:
            print(f"Warning: Could not add employee chart: {e}")
    
    def _add_monthly_comparison_chart(self, workbook: Workbook, trends_df: pd.DataFrame):
        """Add monthly budget vs actual comparison chart."""
        try:
            ws_chart = workbook.create_sheet('Monthly Comparison')
            
            # Write comparison data
            headers = ['Month', 'Budget', 'Actual', 'Variance']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_chart.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
            
            for row_idx, (_, row) in enumerate(trends_df.iterrows(), start=2):
                ws_chart.cell(row=row_idx, column=1, value=row.get('Month', ''))
                ws_chart.cell(row=row_idx, column=2, value=row.get('Total Budget', 0))
                ws_chart.cell(row=row_idx, column=3, value=row.get('Total Actual', 0))
                ws_chart.cell(row=row_idx, column=4, value=row.get('Total Variance', 0))
            
            # Create combined chart (bar + line)
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Monthly Budget vs Actual Comparison"
            chart.y_axis.title = "Amount ($)"
            chart.x_axis.title = "Month"
            
            # Add budget and actual as bars
            data = Reference(ws_chart, min_col=2, min_row=1, max_row=len(trends_df) + 1, max_col=3)
            cats = Reference(ws_chart, min_col=1, min_row=2, max_row=len(trends_df) + 1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws_chart.add_chart(chart, "F2")
            
            # Auto-adjust column widths
            for col in range(1, 5):
                ws_chart.column_dimensions[chr(64 + col)].width = 20
        except Exception as e:
            print(f"Warning: Could not add monthly comparison chart: {e}")
    
    def export_to_csv(self, df: pd.DataFrame, filepath: str):
        """Export DataFrame to CSV file."""
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(filepath, index=False)
        return filepath
    
    def export_to_google_sheets(self, df: pd.DataFrame, spreadsheet_id: Optional[str] = None,
                                sheet_name: str = "VarianceReport"):
        """
        Export DataFrame to Google Sheets.
        
        Args:
            df: DataFrame to export
            spreadsheet_id: Google Sheets spreadsheet ID (uses config if not provided)
            sheet_name: Name of the sheet to create/update (no spaces recommended)
        """
        if not self.sheets_service:
            raise Exception("Google Sheets not initialized. Check credentials.")
        
        spreadsheet_id = spreadsheet_id or settings.google_sheets_spreadsheet_id
        if not spreadsheet_id:
            raise Exception("No spreadsheet ID provided")
        
        try:
            # Convert DataFrame to list of lists
            values = [df.columns.tolist()] + df.values.tolist()
            
            # Check if sheet exists, create if not
            try:
                # Get spreadsheet metadata
                spreadsheet = self.sheets_service.spreadsheets().get(
                    spreadsheetId=spreadsheet_id
                ).execute()
                
                # Check if sheet exists
                sheet_exists = any(
                    sheet['properties']['title'] == sheet_name 
                    for sheet in spreadsheet.get('sheets', [])
                )
                
                if not sheet_exists:
                    # Create new sheet
                    self.sheets_service.spreadsheets().batchUpdate(
                        spreadsheetId=spreadsheet_id,
                        body={
                            'requests': [{
                                'addSheet': {
                                    'properties': {
                                        'title': sheet_name
                                    }
                                }
                            }]
                        }
                    ).execute()
                
                # Clear existing data
                self.sheets_service.spreadsheets().values().clear(
                    spreadsheetId=spreadsheet_id,
                    range=f"'{sheet_name}'!A1:Z1000"
                ).execute()
            except Exception as e:
                print(f"Warning: Could not check/create sheet: {e}")
            
            # Write data
            body = {
                'values': values
            }
            
            result = self.sheets_service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range=f"'{sheet_name}'!A1",
                valueInputOption='RAW',
                body=body
            ).execute()
            
            return result
        except HttpError as error:
            raise Exception(f"Google Sheets API error: {error}")

